from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.action_chains import ActionChains
import traceback
from openpyxl import load_workbook
import time
import datetime
# #Login

class BasePage(object):
    def __init__(self, driver):
        self.driver = driver
# Header Sidebar


class SupervisionalPage(BasePage):
    def headerSidebar(self):
        try:
            return WebDriverWait(self.driver, 120).until(EC.visibility_of_element_located((By.XPATH, "/html/body/div[2]/div[1]/div[1]/div/div/div/div[1]")))
        except NoSuchElementException:
            # If the element was not found, print an error message
            print("Error: Dashboard element not found on page")
        finally:
            # If the element was found, print a success message
            print("Success: Dashboard Element found on page")

# Supervisional BE
# Find SystemProvider BE under Business Entity Management Group
    def findSystemProvider(self):
        systemProvider = "/html/body/div[2]/div/div[1]/div/div/div/section/div[44]/a"
        elementSystemProvider = self.driver.find_element(
            By.XPATH, systemProvider)
        self.driver.execute_script(
            "arguments[0].scrollIntoView(true);", elementSystemProvider)
        try:
            return WebDriverWait(self.driver, 20).until(EC.visibility_of_element_located((By.XPATH, systemProvider)))
        except NoSuchElementException:
            # If the element was not found, print an error message
            print("Error: Element not found on page")
        finally:
            # If the element was found, print a success message
            print("Success: Berhasil Menemukan System Provider")

    def findSupervisionalBE(self):
        # supervisionalBE = "/html/body/div[2]/div/div[1]/div/div/div/section/div[45]/a"
        supervisionalBE = "/html/body/div[2]/div/div[1]/div/div/div/section/div[46]/a"
        element_SupervisionalBE = self.driver.find_element(
            By.XPATH, supervisionalBE)
        self.driver.execute_script(
            "arguments[0].scrollIntoView(true);", element_SupervisionalBE)
        try:
            return WebDriverWait(self.driver, 20).until(EC.visibility_of_element_located((By.XPATH, supervisionalBE)))
        except NoSuchElementException:
            print("Error: Supervisional BE not found")
        finally:
            print("Success: Found Supervisional BE")

    def clickSupervisionalBE(self):
        # supervisionalBE = "/html/body/div[2]/div/div[1]/div/div/div/section/div[45]/a"
        supervisionalBE = "/html/body/div[2]/div/div[1]/div/div/div/section/div[46]/a"
        element_SupervisionalBE = self.driver.find_element(
            By.XPATH, supervisionalBE)
        self.driver.execute_script(
            "arguments[0].scrollIntoView(true);", element_SupervisionalBE)
        element_SupervisionalBE.click()
        print("Elemen Supervisional BE berhasil diklik")

    def findOperationalBE(self):
        operationalBE = "/html[1]/body[1]/div[2]/div[1]/div[1]/div[1]/div[1]/div[1]/section[1]/div[46]/a[1]"
        elementOperationalBE = self.driver.find_element(
            By.XPATH, operationalBE)
        self.driver.execute_script(
            "arguments[0].scrollIntoView(true);", elementOperationalBE)
        try:
            assert WebDriverWait(self.driver, 20).until(
                EC.visibility_of_element_located((By.XPATH, operationalBE)))
        finally:
            print("Berhasil menemukan Operational BE")
        # elementOperationalBE.click()

# #Header Teks Supervisional BE
# def headerSupervisionalBE():
#     element_presentSupervisionalBE = EC.presence_of_all_elements_located((By.CSS_SELECTOR, '#content-area > div.content-wrapper > div.router-view > div > div.router-header.flex.flex-wrap.items-center > div.content-area__heading.pr-4.border-0.md\:border-r.border-t-0.border-b-0.border-l-0.border-solid.border-grey-light.m-4 > h2'))
#     WebDriverWait(browser, tunda).until(element_presentSupervisionalBE)
#     assert element_presentSupervisionalBE.text == 'Supervisional BE'

# #Tombol Add Supervisional BE
    def navigate_and_access_SuperBE(self):
        actions = ActionChains(self.driver)
        actions.move_by_offset(268, 316).perform()
        print("move to inner scrollbar")
        time.sleep(10)
        elementSuperBE = self.driver.find_element(
            By.XPATH, "/html/body/div[2]/div/div[1]/div/div/div/section/div[45]/a")
        self.driver.execute_script(
            "arguments[0].scrollIntoView();", elementSuperBE)
        elementSuperBE.click()

    def tombolAddSupervisionalBE(self):
        addSupervisionalBE = "/html/body/div[2]/div[1]/div[4]/div[2]/div[2]/div/div[2]/div[2]/div[2]/div[1]/div/div[1]/div[2]/div/button"
        try:
            assert WebDriverWait(self.driver, 20).until(
                EC.visibility_of_element_located((By.XPATH, addSupervisionalBE)))
        finally:
            print("Button Add Supervisional BE Found")
        WebDriverWait(self.driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, addSupervisionalBE))).click()
        print("Success Click Button Add Supervisional BE")
    # open workbook

    def excelSVBE(self):
        wb = load_workbook(filename="/Users/macintosh/Documents/SeleUnit/data_supervisional_be.xlsx")
        sheetRange = wb['Sheet1']
        max_row = sheetRange.max_row

        for i in range(2, max_row + 1):
            LegalName = sheetRange['C' + str(i)].value
            BrandName = sheetRange['D' + str(i)].value
            Phone = sheetRange['E' + str(i)].value
            Email = sheetRange['F' + str(i)].value
            WebAddress = sheetRange['G' + str(i)].value
            Address = sheetRange['H' + str(i)].value
            TaxCode = sheetRange['I' + str(i)].value
            Username = sheetRange['J' + str(i)].value
            Password = sheetRange['K' + str(i)].value

            yield LegalName, BrandName, Phone, Email, WebAddress, Address, TaxCode, Username, Password



    def navigate_and_submit_form(self):
            addSupervisionalBE = "/html/body/div[2]/div[1]/div[4]/div[2]/div[2]/div/div[2]/div[2]/div[2]/div[1]/div/div[1]/div[2]/div/button"
            dropdownParent = "/html/body/div[2]/div[1]/div[4]/div[2]/div[2]/div/div[2]/div[2]/div[1]/div[1]/div/div/div[1]/div[2]/div[2]/div/div/div[1]"
            inputDropdownParent = "/html/body/div[2]/div[1]/div[4]/div[2]/div[2]/div/div[2]/div[2]/div[1]/div[1]/div/div/div[1]/div[2]/div[2]/div/div/div[1]/input"
            legal_name = "/html/body/div[2]/div[1]/div[4]/div[2]/div[2]/div/div[2]/div[2]/div[1]/div[1]/div/div/div[1]/div[3]/div[2]/div/div/input"
            brand_name = "/html/body/div[2]/div[1]/div[4]/div[2]/div[2]/div/div[2]/div[2]/div[1]/div[1]/div/div/div[1]/div[4]/div[2]/div/div/input"
            phone_number = "/html/body/div[2]/div[1]/div[4]/div[2]/div[2]/div/div[2]/div[2]/div[1]/div[1]/div/div/div[1]/div[5]/div[2]/div/div/input"
            email_address = "/html/body/div[2]/div[1]/div[4]/div[2]/div[2]/div/div[2]/div[2]/div[1]/div[1]/div/div/div[1]/div[6]/div[2]/div/div/input"
            web_address = "/html/body/div[2]/div[1]/div[4]/div[2]/div[2]/div/div[2]/div[2]/div[1]/div[1]/div/div/div[1]/div[7]/div[2]/div/div/input"
            address_home = "/html/body/div[2]/div[1]/div[4]/div[2]/div[2]/div/div[2]/div[2]/div[1]/div[1]/div/div/div[1]/div[8]/div[2]/div/div/input"
            tax_code = "/html/body/div[2]/div[1]/div[4]/div[2]/div[2]/div/div[2]/div[2]/div[1]/div[1]/div/div/div[1]/div[9]/div[2]/div/div/input"
            credentials_username = "/html/body/div[2]/div[1]/div[4]/div[2]/div[2]/div/div[2]/div[2]/div[1]/div[1]/div/div/div[2]/div[2]/div[2]/div/div/input"
            credentials_password = "/html/body/div[2]/div[1]/div[4]/div[2]/div[2]/div/div[2]/div[2]/div[1]/div[1]/div/div/div[2]/div[3]/div[2]/div/div/input"
            credentials_confirmpassword = "/html/body/div[2]/div[1]/div[4]/div[2]/div[2]/div/div[2]/div[2]/div[1]/div[1]/div/div/div[2]/div[4]/div[2]/div/div/input"
            startDate = "/html/body/div[2]/div[1]/div[4]/div[2]/div[2]/div/div[2]/div[2]/div[2]/div[1]/div/div[2]/div[2]/input"
            inputDateForStartDate = "//div[@class='flatpickr-calendar animate open arrowBottom arrowLeft']//span[@aria-label='July 18, 2023'][normalize-space()='18']"
            buttonSaveAddSupervisionalBE = "/html/body/div[2]/div[1]/div[4]/div[2]/div[2]/div/div[2]/div[2]/div[3]/div[1]/div/div/div[2]/button"

            for LegalName, BrandName, Phone, Email, WebAddress, Address, TaxCode, Username, Password in self.excelSVBE():
                    WebDriverWait(self.driver, 30).until(
                        EC.element_to_be_clickable((By.XPATH, addSupervisionalBE))).click()
                    WebDriverWait(self.driver, 30).until(
                        EC.element_to_be_clickable((By.XPATH, dropdownParent))).click()
                    element_inputDropdownParent = WebDriverWait(self.driver, 10).until(
                        EC.element_to_be_clickable((By.XPATH, inputDropdownParent)))
                    element_inputDropdownParent.send_keys("maximus")
                    WebDriverWait(self.driver, 60).until(EC.text_to_be_present_in_element((By.XPATH, "//*[contains(text(), 'maximus')]"), "maximus"))
                    element_inputDropdownParent.send_keys(Keys.RETURN)
                    legalNameInput = self.driver.find_element(
                        By.XPATH, legal_name)
                    legalNameInput.clear()
                    legalNameInput.send_keys(LegalName)

                    brandNameInput = self.driver.find_element(
                        By.XPATH, brand_name)
                    brandNameInput.clear()
                    brandNameInput.send_keys(BrandName)

                    phoneInput = self.driver.find_element(
                        By.XPATH, phone_number)
                    phoneInput.clear()
                    phoneInput.send_keys(Phone)

                    emailInput = self.driver.find_element(
                        By.XPATH, email_address)
                    emailInput.clear()
                    emailInput.send_keys(Email)

                    webAddressInput = self.driver.find_element(
                        By.XPATH, web_address)
                    webAddressInput.clear()
                    webAddressInput.send_keys(WebAddress)

                    addressInput = self.driver.find_element(
                        By.XPATH, address_home)
                    addressInput.clear()
                    addressInput.send_keys(Address)

                    taxCodeInput = self.driver.find_element(
                        By.XPATH, tax_code)
                    taxCodeInput.clear()
                    taxCodeInput.send_keys(TaxCode)

                    usernameInput = self.driver.find_element(
                        By.XPATH, credentials_username)
                    usernameInput.clear()
                    usernameInput.send_keys(Username)

                    passwordInput = self.driver.find_element(
                        By.XPATH, credentials_password)
                    passwordInput.clear()
                    passwordInput.send_keys(Password)

                    confirmPasswordInput = self.driver.find_element(
                        By.XPATH, credentials_confirmpassword)
                    confirmPasswordInput.clear()
                    confirmPasswordInput.send_keys(Password)

                    WebDriverWait(self.driver, 10).until(EC.element_to_be_clickable((By.XPATH, startDate))).click()
                    print("Berhasil klik Kolom untuk Kalender Start Date")
                    WebDriverWait(self.driver, 20).until(EC.element_to_be_clickable((By.XPATH, inputDateForStartDate))).click()
                    buttonSaveAddSupervisionalBE = "//span[normalize-space()='Submit']"
                    WebDriverWait(self.driver, 20).until(EC.element_to_be_clickable((By.XPATH, buttonSaveAddSupervisionalBE))).click()
                    if not self.detectAlert() :
                        continue
                    elif not self.detectInternalServerError():
                        continue

    def detectInternalServerError(self):
        errorIcon = "//span[@class='swal2-x-mark']"
        errorTitle = "//h2[@id='swal2-title']"
        errorBody = "//div[@id='swal2-content']"
        buttonOk = "//button[normalize-space()='OK']"

        # Check for error
        try:
            error_title = WebDriverWait(self.driver, 20).until(EC.visibility_of_element_located((By.XPATH, errorTitle)))
            error_body = WebDriverWait(self.driver, 20).until(EC.visibility_of_element_located((By.XPATH, errorBody)))

            if WebDriverWait(self.driver, 20).until(EC.visibility_of_element_located((By.XPATH, errorIcon))):
                title_text = error_title.text
                body_text = error_body.text
                traceback.print_exc()
                timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                error_code = "Internal Server Error"  # Replace with the actual error type
                page_url = self.driver.current_url
        
                error_text = f"Timestamp: {timestamp}\nError Code: {error_code}\nPage URL: {page_url}\nTitle: {title_text}\nBody: {body_text}"
                with open("result.txt", "a") as file:
                    file.write("\n")
                    file.write(error_text)
                WebDriverWait(self.driver, 10).until(EC.element_to_be_clickable((By.XPATH, buttonOk))).click()
                self.driver.back()
                
                # return True
        except TimeoutException:
            print("Error: Element error not found on page")
            
            # return False
        except Exception as e:
            traceback.print_exc()
            timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            error_code = str(type(e).__name__)
            page_url = self.driver.current_url

            error_text = f"Timestamp: {timestamp}\nError Code: {error_code}\nPage URL: {page_url}\nTitle: {title_text}\nBody: {body_text}"

            with open("result.txt", "a") as file:
                file.write("\n")
                file.write("\n")
                file.write(error_text)

            # return False
        
    def detectAlert(self):
            failedAlert = "//div[@class='vs-component vs-notifications vs-noti-top-right vs-noti-warning activeNoti']"
            failedAlertTitle = "/html/body/div[1]/div/div/h3"
            failedAlertBody = "/html/body/div[1]/div/div/p"

            # Check for error
            try:
                error_alertTitle = WebDriverWait(self.driver, 20).until(EC.visibility_of_element_located((By.XPATH, failedAlertTitle)))
                error_alertBody = WebDriverWait(self.driver, 20).until(EC.visibility_of_element_located((By.XPATH, failedAlertBody)))
                # Set up an explicit wait to wait for the alert error element to be present or visible
                wait = WebDriverWait(self.driver, 10)
                if WebDriverWait(self.driver, 20).until(EC.visibility_of_element_located((By.XPATH, failedAlertTitle))):
                    alertTitle_text = error_alertTitle.text
                    alertBody_text = error_alertBody.text
                    traceback.print_exc()
                    timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    error_code = "Failed Alert"  # Replace with the actual error type
                    page_url = self.driver.current_url
            
                    error_text = f"Timestamp: {timestamp}\nError Code: {error_code}\nPage URL: {page_url}\nTitle: {alertTitle_text}\nBody: {alertBody_text}"
                    with open("result.txt", "a") as file:
                        file.write("\n")
                        file.write("\n")
                        file.write(error_text)
                    self.driver.back()
                    # return True
                
            except TimeoutException:
                print("Error: Element error not found on page")
                # return False
            except Exception as e:
                traceback.print_exc()
                timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                error_code = str(type(e).__name__)
                page_url = self.driver.current_url

                error_text = f"Timestamp: {timestamp}\nError Code: {error_code}\nPage URL: {page_url}\nTitle: {alertTitle_text}\nBody: {alertBody_text}"

                with open("result.txt", "a") as file:
                    file.write("\n")
                    file.write(error_text)
                # return False

    # def navigate_and_submit_form(self):
    #     addSupervisionalBE = "/html/body/div[2]/div[1]/div[4]/div[2]/div[2]/div/div[2]/div[2]/div[2]/div[1]/div/div[1]/div[2]/div/button"
    #     dropdownParent = "/html/body/div[2]/div[1]/div[4]/div[2]/div[2]/div/div[2]/div[2]/div[1]/div[1]/div/div/div[1]/div[2]/div[2]/div/div/div[1]"
    #     inputDropdownParent = "/html/body/div[2]/div[1]/div[4]/div[2]/div[2]/div/div[2]/div[2]/div[1]/div[1]/div/div/div[1]/div[2]/div[2]/div/div/div[1]/input"
    #     legal_name = "/html/body/div[2]/div[1]/div[4]/div[2]/div[2]/div/div[2]/div[2]/div[1]/div[1]/div/div/div[1]/div[3]/div[2]/div/div/input"
    #     brand_name = "/html/body/div[2]/div[1]/div[4]/div[2]/div[2]/div/div[2]/div[2]/div[1]/div[1]/div/div/div[1]/div[4]/div[2]/div/div/input"
    #     phone_number = "/html/body/div[2]/div[1]/div[4]/div[2]/div[2]/div/div[2]/div[2]/div[1]/div[1]/div/div/div[1]/div[5]/div[2]/div/div/input"
    #     email_address = "/html/body/div[2]/div[1]/div[4]/div[2]/div[2]/div/div[2]/div[2]/div[1]/div[1]/div/div/div[1]/div[6]/div[2]/div/div/input"
    #     web_address = "/html/body/div[2]/div[1]/div[4]/div[2]/div[2]/div/div[2]/div[2]/div[1]/div[1]/div/div/div[1]/div[7]/div[2]/div/div/input"
    #     address_home = "/html/body/div[2]/div[1]/div[4]/div[2]/div[2]/div/div[2]/div[2]/div[1]/div[1]/div/div/div[1]/div[8]/div[2]/div/div/input"
    #     tax_code = "/html/body/div[2]/div[1]/div[4]/div[2]/div[2]/div/div[2]/div[2]/div[1]/div[1]/div/div/div[1]/div[9]/div[2]/div/div/input"
    #     credentials_username = "/html/body/div[2]/div[1]/div[4]/div[2]/div[2]/div/div[2]/div[2]/div[1]/div[1]/div/div/div[2]/div[2]/div[2]/div/div/input"
    #     credentials_password = "/html/body/div[2]/div[1]/div[4]/div[2]/div[2]/div/div[2]/div[2]/div[1]/div[1]/div/div/div[2]/div[3]/div[2]/div/div/input"
    #     credentials_confirmpassword = "/html/body/div[2]/div[1]/div[4]/div[2]/div[2]/div/div[2]/div[2]/div[1]/div[1]/div/div/div[2]/div[4]/div[2]/div/div/input"
    #     startDate = "/html/body/div[2]/div[1]/div[4]/div[2]/div[2]/div/div[2]/div[2]/div[2]/div[1]/div/div[2]/div[2]/input"
    #     # inputDateForStartDate = "//div[@class='flatpickr-calendar animate open arrowBottom arrowLeft']//span[@aria-label='Jun 15, 2022'][normalize-space()='14']"
    #     inputDateForStartDate = "//div[@class='flatpickr-calendar animate open arrowBottom arrowLeft']//span[@aria-label='June 13, 2023'][normalize-space()='13']"
        
    #     buttonSaveAddSupervisionalBE = "/html/body/div[2]/div[1]/div[4]/div[2]/div[2]/div/div[2]/div[2]/div[3]/div[1]/div/div/div[2]/button"

    #     for LegalName, BrandName, Phone, Email, WebAddress, Address, TaxCode, Username, Password in self.excelSVBE():
    #         WebDriverWait(self.driver, 10).until(
    #             EC.element_to_be_clickable((By.XPATH, addSupervisionalBE))).click()
    #         WebDriverWait(self.driver, 10).until(
    #             EC.element_to_be_clickable((By.XPATH, dropdownParent))).click()
    #         element_inputDropdownParent = WebDriverWait(self.driver, 10).until(
    #             EC.element_to_be_clickable((By.XPATH, inputDropdownParent)))
    #         element_inputDropdownParent.send_keys("maximus")
    #         time.sleep(3)
    #         print("Berhasil select dropdown")
    #         element_inputDropdownParent.send_keys(Keys.RETURN)
    #         print("Berhasil Select Dropdown pada Form Parent")  
    #         # Fill the form inputs
    #         # (Assuming you have appropriate locators and elements for each form input field)
    #         legalNameInput = self.driver.find_element(
    #             By.XPATH, legal_name)
    #         legalNameInput.clear()
    #         legalNameInput.send_keys(LegalName)

    #         brandNameInput = self.driver.find_element(
    #             By.XPATH, brand_name)
    #         brandNameInput.clear()
    #         brandNameInput.send_keys(BrandName)

    #         phoneInput = self.driver.find_element(
    #             By.XPATH, phone_number)
    #         phoneInput.clear()
    #         phoneInput.send_keys(Phone)

    #         emailInput = self.driver.find_element(
    #             By.XPATH, email_address)
    #         emailInput.clear()
    #         emailInput.send_keys(Email)

    #         webAddressInput = self.driver.find_element(
    #             By.XPATH, web_address)
    #         webAddressInput.clear()
    #         webAddressInput.send_keys(WebAddress)

    #         addressInput = self.driver.find_element(
    #             By.XPATH, address_home)
    #         addressInput.clear()
    #         addressInput.send_keys(Address)

    #         taxCodeInput = self.driver.find_element(
    #             By.XPATH, tax_code)
    #         taxCodeInput.clear()
    #         taxCodeInput.send_keys(TaxCode)

    #         usernameInput = self.driver.find_element(
    #             By.XPATH, credentials_username)
    #         usernameInput.clear()
    #         usernameInput.send_keys(Username)

    #         passwordInput = self.driver.find_element(
    #             By.XPATH, credentials_password)
    #         passwordInput.clear()
    #         passwordInput.send_keys(Password)

    #         confirmPasswordInput = self.driver.find_element(
    #             By.XPATH, credentials_confirmpassword)
    #         confirmPasswordInput.clear()
    #         confirmPasswordInput.send_keys(Password)

    #         WebDriverWait(self.driver, 10).until(EC.element_to_be_clickable((By.XPATH, startDate))).click()
    #         print("Berhasil klik Kolom untuk Kalender Start Date")
    #         WebDriverWait(self.driver, 20).until(EC.element_to_be_clickable((By.XPATH, inputDateForStartDate))).click()
    #         buttonSaveAddSupervisionalBE = "//span[normalize-space()='Submit']"
    #         WebDriverWait(self.driver, 20).until(EC.element_to_be_clickable((By.XPATH, buttonSaveAddSupervisionalBE))).click()
    #         # print("Berhasil menyimpan Add Supervisional BE")
    #         # Check for internal server error
    #         if not self.detectInternalServerError():
    #             # If error not detected, continue with the next iteration
    #             continue

    # def detectInternalServerError(self):
    #     errorIcon = "//span[@class='swal2-x-mark']"
    #     errorTitleMessage = "//h2[@id='swal2-title']"
    #     errorBodyMessage = "//div[@id='swal2-content']"
    #     buttonOk = "//button[normalize-space()='OK']"
        
    #     try:
    #         # Wait for the error icon to be visible
    #         WebDriverWait(self.driver, 20).until(EC.visibility_of_element_located((By.XPATH, errorIcon)))
            
    #         # Get the text of the error title and body messages
    #         title_text = self.driver.find_element(By.XPATH, errorTitleMessage).text
    #         body_text = self.driver.find_element(By.XPATH, errorBodyMessage).text
            
    #         # Perform assertions on the error messages
    #         assert "Error occurred" in title_text
    #         assert "Error occurred" in body_text
            
    #         # Click the OK button
    #         self.driver.find_element(By.XPATH, buttonOk).click()
            
    #         # Return True to indicate the error was detected and handled
    #         return True
    
    #     except TimeoutException:
    #         # If the element was not found, print an error message
    #         print("Error: Element error not found on page")
    #         return False
    #     except Exception as e:
    #         # If any other exception occurs, print the traceback and write the error text to a file
    #         traceback.print_exc()
    #         error_text = f"Error Text: {title_text} - {body_text}"
    #         with open("result.txt", "w") as file:
    #             file.write(error_text)
    #         return False
        
    def formParent(self):
        dropdownParent = "/html/body/div[2]/div[1]/div[4]/div[2]/div[2]/div/div[2]/div[2]/div[1]/div[1]/div/div/div[1]/div[2]/div[2]/div/div/div[1]"
        inputDropdownParent = "/html/body/div[2]/div[1]/div[4]/div[2]/div[2]/div/div[2]/div[2]/div[1]/div[1]/div/div/div[1]/div[2]/div[2]/div/div/div[1]/input"

        WebDriverWait(self.driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, dropdownParent))).click()
        element_inputDropdownParent = WebDriverWait(self.driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, inputDropdownParent)))
        element_inputDropdownParent.send_keys("maximus")
        time.sleep(10)
        print("Berhasil select dropdown")
        element_inputDropdownParent.send_keys(Keys.RETURN)
        print("Berhasil Select Dropdown pada Form Parent")

    # #Form Legal Name

    def formLegalName(self):
        legal_name = "/html/body/div[2]/div[1]/div[4]/div[2]/div[2]/div/div[2]/div[2]/div[1]/div[1]/div/div/div[1]/div[3]/div[2]/div/div/input"
        brand_name = "/html/body/div[2]/div[1]/div[4]/div[2]/div[2]/div/div[2]/div[2]/div[1]/div[1]/div/div/div[1]/div[4]/div[2]/div/div/input"
        WebDriverWait(self.driver, 10).until(EC.element_to_be_clickable(
            (By.XPATH, legal_name))).send_keys("Selenium_Unit_Test")
        print("Berhasil memasukkan form nama")

# #Move Tab
# def moveTab():
#     legal_name = "/html/body/div[2]/div[1]/div[4]/div[2]/div[2]/div/div[2]/div[2]/div[1]/div[1]/div/div/div[1]/div[3]/div[2]/div/div/input"
#     moveTab = WebDriverWait(browser, 10).until(EC.element_to_be_clickable((By.XPATH, legal_name)))
#     moveTab.click()
#     moveTab.send_keys(Keys.TAB)

# #Form Brand Name
# def formBrandName():
#     brand_name = "/html/body/div[2]/div[1]/div[4]/div[2]/div[2]/div/div[2]/div[2]/div[1]/div[1]/div/div/div[1]/div[4]/div[2]/div/div/input"
#     WebDriverWait(browser, 10).until(EC.element_to_be_clickable((By.XPATH, brand_name))).send_keys("Brand Testd10")
#     print("Berhasil memasukkan form Brand Name")

# #Form Phone Number
# def formPhoneNumber():
#     phone_number = "/html/body/div[2]/div[1]/div[4]/div[2]/div[2]/div/div[2]/div[2]/div[1]/div[1]/div/div/div[1]/div[5]/div[2]/div/div/input"
#     WebDriverWait(browser, 10).until(EC.element_to_be_clickable((By.XPATH, phone_number))).send_keys("081aa///")
#     print("Berhasil memasukkan form Phone Number")

# #Form Email
# def formEmail():
#     email_address = "/html/body/div[2]/div[1]/div[4]/div[2]/div[2]/div/div[2]/div[2]/div[1]/div[1]/div/div/div[1]/div[6]/div[2]/div/div/input"
#     WebDriverWait(browser, 10).until(EC.element_to_be_clickable((By.XPATH, email_address))).send_keys("deltatest10@gmail.com")
#     print("Berhasil memasukkan form email")

# #Form Web Address
# def formWebAddress():
#     web_address = "/html/body/div[2]/div[1]/div[4]/div[2]/div[2]/div/div[2]/div[2]/div[1]/div[1]/div/div/div[1]/div[7]/div[2]/div/div/input"
#     WebDriverWait(browser, 10).until(EC.element_to_be_clickable((By.XPATH, web_address))).send_keys("deltatest10.blogspot.com")
#     print("Berhasil memasukkan Web Address")

# #Form Address Home
# def formAddressHome():
#     address_home = "/html/body/div[2]/div[1]/div[4]/div[2]/div[2]/div/div[2]/div[2]/div[1]/div[1]/div/div/div[1]/div[8]/div[2]/div/div/input"
#     WebDriverWait(browser, 10).until(EC.element_to_be_clickable((By.XPATH, address_home))).send_keys("Test 10")
#     print("Berhasil memasukkan form Address Home")

# #Form Tax Code
# def formTaxCode():
#     tax_code = "/html/body/div[2]/div[1]/div[4]/div[2]/div[2]/div/div[2]/div[2]/div[1]/div[1]/div/div/div[1]/div[9]/div[2]/div/div/input"
#     WebDriverWait(browser, 10).until(EC.element_to_be_clickable((By.XPATH, tax_code))).send_keys("557210")
#     print("Berhasil memasukkan form Tax Code")

# #Form username credentials
# def usernameCredentials():
#     credentials_username = "/html/body/div[2]/div[1]/div[4]/div[2]/div[2]/div/div[2]/div[2]/div[1]/div[1]/div/div/div[2]/div[2]/div[2]/div/div/input"
#     WebDriverWait(browser, 10).until(EC.element_to_be_clickable((By.XPATH, credentials_username))).send_keys("account_testd10")
#     print("Berhasil memasukkan Username pada Form Credentials")

# #Form password credentials
# def passwordCredentials():
#     credentials_password = "/html/body/div[2]/div[1]/div[4]/div[2]/div[2]/div/div[2]/div[2]/div[1]/div[1]/div/div/div[2]/div[3]/div[2]/div/div/input"
#     WebDriverWait(browser, 10).until(EC.element_to_be_clickable((By.XPATH, credentials_password))).send_keys("Merchant123")
#     print("Berhasil memasukkan password credentials")

# def eyePasswordCredentials():
#     eyePassword = "/html/body/div[2]/div/div[4]/div[2]/div[2]/div/div[2]/div[2]/div[1]/div[1]/div/div/div[2]/div[3]/div[3]/button"
#     WebDriverWait(browser, 10).until(EC.element_to_be_clickable((By.XPATH, eyePassword))).click()
#     print("Berhasil mengaktifkan eye pada password")

# def confirmPasswordCredentials():
#     credentials_confirmpassword = "/html/body/div[2]/div[1]/div[4]/div[2]/div[2]/div/div[2]/div[2]/div[1]/div[1]/div/div/div[2]/div[4]/div[2]/div/div/input"
#     WebDriverWait(browser, 10).until(EC.element_to_be_clickable((By.XPATH, credentials_confirmpassword))).send_keys("Merchant123")
#     print("Berhasil memasukkan confirm password credentials")

# #Form Contract
# def contractStart():
#     startDate = "/html/body/div[2]/div[1]/div[4]/div[2]/div[2]/div/div[2]/div[2]/div[2]/div[1]/div/div[2]/div[2]/input"
#     inputDateForStartDate = "//div[@class='flatpickr-calendar animate open arrowBottom arrowLeft']//span[@aria-label='September 14, 2022'][normalize-space()='14']"
#     WebDriverWait(browser, 10).until(EC.element_to_be_clickable((By.XPATH, startDate))).click()
#     print("Berhasil klik Kolom untuk Kalender Start Date")
#     WebDriverWait(browser, 20).until(EC.element_to_be_clickable((By.XPATH, inputDateForStartDate))).click()
#     print("Berhasil input tanggal pada Start Date")

# def contractEnd():
#     endDate = "/html/body/div[2]/div[1]/div[4]/div[2]/div[2]/div/div[2]/div[2]/div[2]/div[1]/div/div[3]/div[2]/input"
#     nextMonth = "//div[@class='flatpickr-calendar animate open arrowBottom arrowLeft']//span[@class='flatpickr-next-month']//*[name()='svg']"
#     inputDateForEndDate = "//span[@aria-label='October 17, 2022']"
#     WebDriverWait(browser, 10).until(EC.element_to_be_clickable((By.XPATH, endDate))).click()
#     print("Berhasil klik Kolom untuk Kalender End Date")
#     WebDriverWait(browser, 20).until(EC.element_to_be_clickable((By.XPATH, nextMonth))).click()
#     print("Berhasil melakukan geser bulan")
#     WebDriverWait(browser, 20).until(EC.element_to_be_clickable((By.XPATH, inputDateForEndDate))).click()
#     print("Berhasil input tanggal pada End Date")

# #Button Save Add Supervisional BE
# def SaveAddSupervisionalBE():
#     buttonSaveAddSupervisionalBE = "/html/body/div[2]/div[1]/div[4]/div[2]/div[2]/div/div[2]/div[2]/div[3]/div[1]/div/div/div[2]/button"
#     WebDriverWait(browser, 20).until(EC.element_to_be_clickable((By.XPATH, buttonSaveAddSupervisionalBE))).click()
#     print("Berhasil menyimpan Add Supervisional BE")

# #Button Cancel Add Supervisional BE
# def CancelAddSupervisionalBE():
#     cancel = "/html/body/div[2]/div[1]/div[4]/div[2]/div[2]/div/div[2]/div[2]/div[3]/div[1]/div/div/div[1]/button"
#     WebDriverWait(browser, 20).until(EC.element_to_be_clickable((By.XPATH, cancel))).click()
#     print("Berhasil Cancel Add Supervisional BE")

# #Button Language
# def buttonLanguage():
#     clickLanguage = "//div[@class='flex justify-content-end']//button[@type='button']"
#     chooseVietnamLanguage = "/html/body/div[1]/div[1]/li[2]/a"
#     WebDriverWait(browser, 20).until(EC.element_to_be_clickable((By.XPATH, clickLanguage))).click()
#     print("Berhasil Click Button Language Language")
#     WebDriverWait(browser, 20).until(EC.element_to_be_clickable((By.XPATH, chooseVietnamLanguage))).click()
#     print("Berhasil mengubah ke bahasa Vietnam")

# #Button Edit Supervisional BE
# def buttonEditSupervisionalBE():
#     editsupervisionalBE = "/html/body/div[2]/div[1]/div[4]/div[2]/div[2]/div/div[2]/div[2]/div[2]/div[1]/div/div[2]/div/div[1]/div/div[1]/div/div[3]/div[3]/div[7]/div/div/button[1]"
#     WebDriverWait(browser, 20).until(EC.element_to_be_clickable((By.XPATH, editsupervisionalBE))).click()
#     print("Berhasil klik edit supervisional BE")

# #Form Edit Parent
# def formEditParent():
#     dropdownEditParent = "/html/body/div[2]/div[1]/div[4]/div[2]/div[2]/div/div[2]/div[2]/div[1]/div[1]/div/div/div[1]/div[2]/div[2]/div/div"
#     inputDropdownEditParent = "/html/body/div[2]/div[1]/div[4]/div[2]/div[2]/div/div[2]/div[2]/div[1]/div[1]/div/div/div[1]/div[2]/div[2]/div/div/div[1]/input"
#     WebDriverWait(browser, 10).until(EC.element_to_be_clickable((By.XPATH, dropdownEditParent))).click()
#     element_editDropdownParent = WebDriverWait(browser, 10).until(EC.element_to_be_clickable((By.XPATH, inputDropdownEditParent)))
#     element_editDropdownParent.send_keys("testsele")
#     element_editDropdownParent.send_keys(Keys.RETURN)
#     print("Berhasil  Edit Form Parent")

# #Form Edit Legal Name
# def formEditLegalName():
#     editLegalName = "/html/body/div[2]/div[1]/div[4]/div[2]/div[2]/div/div[2]/div[2]/div[1]/div[1]/div/div/div[1]/div[3]/div[2]/div/div/input"
#     elementEditLegalName = WebDriverWait(browser, 10).until(EC.element_to_be_clickable((By.XPATH, editLegalName)))
#     elementEditLegalName.send_keys(Keys.CONTROL + "a")
#     elementEditLegalName.send_keys(Keys.BACK_SPACE)
#     elementEditLegalName.send_keys("Edit Testd7")
#     print("Berhasil mengubah form legal name")

# #Form Edit Brand Name
# def formEditBrandName():
#     editBrandName = "/html/body/div[2]/div[1]/div[4]/div[2]/div[2]/div/div[2]/div[2]/div[1]/div[1]/div/div/div[1]/div[4]/div[2]/div/div/input"
#     elementEditBrandName = WebDriverWait(browser, 10).until(EC.element_to_be_clickable((By.XPATH, editBrandName)))
#     elementEditBrandName.send_keys(Keys.CONTROL + "a")
#     elementEditBrandName.send_keys(Keys.BACK_SPACE)
#     elementEditBrandName.send_keys("Edit Brand Testd7")
#     print("Berhasil mengubah form Brand Name")

# #Form Edit Phone Number
# def formEditPhoneNumber():
#     editPhoneNumber = "/html/body/div[2]/div[1]/div[4]/div[2]/div[2]/div/div[2]/div[2]/div[1]/div[1]/div/div/div[1]/div[5]/div[2]/div/div/input"
#     elementEditPhoneNumber = WebDriverWait(browser, 10).until(EC.element_to_be_clickable((By.XPATH, editPhoneNumber)))
#     elementEditPhoneNumber.send_keys(Keys.CONTROL + "a")
#     elementEditPhoneNumber.send_keys(Keys.BACK_SPACE)
#     elementEditPhoneNumber.send_keys("081305111222")
#     print("Berhasil mengubah form Phone Number")

# #Form Edit Email Address
# def formEditEmailAddress():
#     editEmailAddress = "/html/body/div[2]/div[1]/div[4]/div[2]/div[2]/div/div[2]/div[2]/div[1]/div[1]/div/div/div[1]/div[6]/div[2]/div/div/input"
#     elementEditEmailAddress = WebDriverWait(browser, 10).until(EC.element_to_be_clickable((By.XPATH, editEmailAddress)))
#     elementEditEmailAddress.send_keys(Keys.CONTROL + "a")
#     elementEditEmailAddress.send_keys(Keys.BACK_SPACE)
#     elementEditEmailAddress.send_keys("editdeltatest7@gmail.com")
#     print("Berhasil mengubah form email")

# #Form Edit Web Address
# def formEditWebAddress():
#     editWebAddress = "/html/body/div[2]/div[1]/div[4]/div[2]/div[2]/div/div[2]/div[2]/div[1]/div[1]/div/div/div[1]/div[7]/div[2]/div/div/input"
#     elementEditWebAddress = WebDriverWait(browser, 10).until(EC.element_to_be_clickable((By.XPATH, editWebAddress)))
#     elementEditWebAddress.send_keys(Keys.CONTROL + "a")
#     elementEditWebAddress.send_keys(Keys.BACK_SPACE)
#     elementEditWebAddress.send_keys("editdeltatest7.blogspot.com")
#     print("Berhasil mengubah Web Address")

# #Form Edit Address Home
# def formEditAddressHome():
#     editAddressHome = "/html/body/div[2]/div[1]/div[4]/div[2]/div[2]/div/div[2]/div[2]/div[1]/div[1]/div/div/div[1]/div[8]/div[2]/div/div/input"
#     elementEditAddressHome = WebDriverWait(browser, 10).until(EC.element_to_be_clickable((By.XPATH, editAddressHome)))
#     elementEditAddressHome.send_keys(Keys.CONTROL + "a")
#     elementEditAddressHome.send_keys(Keys.BACK_SPACE)
#     elementEditAddressHome.send_keys("Edit Address Test 7")
#     print("Berhasil mengubah form Address Home")

# #Form Edit Tax Code
# def formEditTaxCode():
#     editTaxCode = "/html/body/div[2]/div[1]/div[4]/div[2]/div[2]/div/div[2]/div[2]/div[1]/div[1]/div/div/div[1]/div[9]/div[2]/div/div/input"
#     elementEditTaxCode = WebDriverWait(browser, 10).until(EC.element_to_be_clickable((By.XPATH, editTaxCode)))
#     elementEditTaxCode.send_keys(Keys.CONTROL + "a")
#     elementEditTaxCode.send_keys(Keys.BACK_SPACE)
#     elementEditTaxCode.send_keys("15572111")
#     print("Berhasil mengubah form Tax Code")

# #Form Edit Credentials
# def formEditUsernameCredentials():
#     editUsernameCredentials = "/html/body/div[2]/div[1]/div[4]/div[2]/div[2]/div/div[2]/div[2]/div[1]/div[1]/div/div/div[2]/div[2]/div[2]/div/div/input"
#     elementEditUsernameCredentials = WebDriverWait(browser, 10).until(EC.element_to_be_clickable((By.XPATH, editUsernameCredentials)))
#     elementEditUsernameCredentials.send_keys(Keys.CONTROL + "a")
#     elementEditUsernameCredentials.send_keys(Keys.BACK_SPACE)
#     elementEditUsernameCredentials.send_keys("editaccount_testd7")
#     print("Berhasil mengubah Username pada Form Credentials")

# #Form Edit Contract
# def editContractStart():
#     editStartDate = "/html/body/div[2]/div[1]/div[4]/div[2]/div[2]/div/div[2]/div[2]/div[2]/div[1]/div/div[2]/div[2]/input"
#     editInputDateForStartDate = "//span[@aria-label='September 18, 2022']"
#     WebDriverWait(browser, 10).until(EC.element_to_be_clickable((By.XPATH, editStartDate))).click()
#     print("Berhasil klik Kolom untuk Kalender Start Date")
#     WebDriverWait(browser, 20).until(EC.element_to_be_clickable((By.XPATH, editInputDateForStartDate))).click()
#     print("Berhasil mengubah tanggal pada Start Date")
#     WebDriverWait(browser, 10).until(EC.element_to_be_clickable((By.XPATH, editEndDate))).click()
#     print("Berhasil klik Kolom untuk Kalender End Date")
#     WebDriverWait(browser, 20).until(EC.element_to_be_clickable((By.XPATH, editInputDateForEndDate))).click()
#     print("Berhasil mengubah tanggal pada End Date")

# def editContractEnd():
#     editEndDate = "/html/body/div[2]/div[1]/div[4]/div[2]/div[2]/div/div[2]/div[2]/div[2]/div[1]/div/div[3]/div[2]/input"
#     editInputDateForEndDate = "//span[@aria-label='October 21, 2022']"
#     # editNextMonth = "//div[@class='flatpickr-calendar animate open arrowBottom arrowLeft']//span[@class='flatpickr-next-month']//*[name()='svg']"
#     WebDriverWait(browser, 10).until(EC.element_to_be_clickable((By.XPATH, editEndDate))).click()
#     print("Berhasil klik Kolom untuk Kalender End Date")
#     WebDriverWait(browser, 20).until(EC.element_to_be_clickable((By.XPATH, editInputDateForEndDate))).click()
#     print("Berhasil mengubah tanggal pada End Date")

# #Button View Supervisional BE
# def buttonViewSupervisionalBE():
#     viewSupervisionalBE = "/html/body/div[2]/div[1]/div[4]/div[2]/div[2]/div/div[2]/div[2]/div[2]/div[1]/div/div[2]/div/div[1]/div/div[1]/div/div[3]/div[3]/div[1]/div/div/button[2]"
#     WebDriverWait(browser, 20).until(EC.element_to_be_clickable((By.XPATH, viewSupervisionalBE))).click()
#     print("Berhasil klik view supervisional BE")

# def buttonBackToListSupervisionalBE():
#     backToListSupervisionalBE = "#content-area > div.content-wrapper > div.router-view > div > div.content-area__content > div:nth-child(2) > div:nth-child(3) > div.vx-card__collapsible-content.vs-con-loading__container > div > div > div > button"
#     WebDriverWait(browser, 20).until(EC.element_to_be_clickable((By.CSS_SELECTOR, backToListSupervisionalBE))).click()
#     print("Berhasil kembali ke List Supervisional BE")

# #Button Delete Supervisional BE
# def buttondeleteSupervisionalBE():
#     deleteSupervisionalBE = "/html/body/div[2]/div/div[4]/div[2]/div[2]/div/div[2]/div[2]/div[2]/div[1]/div/div[2]/div/div[1]/div/div[1]/div/div[3]/div[3]/div[9]/div/div/button[3]"
#     WebDriverWait(browser, 20).until(EC.element_to_be_clickable((By.XPATH, deleteSupervisionalBE))).click()
#     print("Berhasil klik delete supervisional BE")

# #Button Cancel Delete Supervisional BE
# def buttonCancelDeleteSupervisionalBE():
#     cancelDeleteSupervisionalBE = "/html/body/div[1]/div[2]/footer/button[2]"
#     WebDriverWait(browser, 20).until(EC.element_to_be_clickable((By.XPATH, cancelDeleteSupervisionalBE))).click()
#     print("Berhasil klik cancel delete supervisional BE")

# #Button Accept Delete Supervisional BE
# def buttonAcceptDeleteSupervisionalBE():
#     acceptDeleteSupervisionalBE = "/html/body/div[1]/div[2]/footer/button[1]"
#     WebDriverWait(browser, 20).until(EC.element_to_be_clickable((By.XPATH, acceptDeleteSupervisionalBE))).click()
#     print("Berhasil klik accept delete supervisional BE")

# #Button Breadcrumb Home
# def buttonBreadCrumbHome():
#     breadCrumbHome = "/html/body/div[2]/div[1]/div[4]/div[2]/div[2]/div/div[1]/div[2]/ul/li[1]/a"
#     WebDriverWait(browser, 20).until(EC.element_to_be_clickable((By.XPATH, breadCrumbHome))).click()
#     print("Berhasil klik BreadCrumbHome")

# #Button Next Page
# def buttonNextPage():
#     nextPage = "/html/body/div[2]/div[1]/div[4]/div[2]/div[2]/div/div[2]/div[2]/div[2]/div[1]/div/div[2]/div/div[2]/nav/button[2]"
#     WebDriverWait(browser, 20).until(EC.element_to_be_clickable((By.XPATH, nextPage))).click()
#     print("Berhasil klik Next Page")

# #Button Previous Page
# def buttonPreviousPage():
#     previousPage = "/html/body/div[2]/div[1]/div[4]/div[2]/div[2]/div/div[2]/div[2]/div[2]/div[1]/div/div[2]/div/div[2]/nav/button[1]"
#     WebDriverWait(browser, 20).until(EC.element_to_be_clickable((By.XPATH, previousPage))).click()
#     print("Berhasil klik Previous Page")

# def formParentfilterInputParent():
#     clickFormParent="/html/body/div[2]/div/div[4]/div[2]/div[2]/div/div[2]/div[2]/div[1]/div[1]/div/div[1]/div[2]/div/div[2]"
#     formInputParent="/html/body/div[2]/div/div[4]/div[2]/div[2]/div/div[2]/div[2]/div[1]/div[1]/div/div[1]/div[2]/div/div[2]/input"
#     WebDriverWait(browser, 20).until(EC.element_to_be_clickable((By.XPATH, clickFormParent))).click()
#     clickFormInputParent = WebDriverWait(browser, 20).until(EC.element_to_be_clickable((By.XPATH, formInputParent)))
#     clickFormInputParent.send_keys("testime")
#     clickFormInputParent.send_keys(Keys.RETURN)
#     print("Berhasil input Parent")

# def buttonFilter():
#     filterButton = "/html/body/div[2]/div/div[4]/div[2]/div[2]/div/div[2]/div[2]/div[1]/div[1]/div/div[5]/div[2]/button[2]"
#     WebDriverWait(browser, 20).until(EC.element_to_be_clickable((By.XPATH, filterButton))).click()
#     print("Berhasil klik filter")

# #Memastikan field tulisannya sama
# def assertField():
#     placeholderLabel = WebDriverWait(browser, 20).until(EC.visibility_of_element_located((By.XPATH, "/html/body/div[2]/div/div[4]/div[2]/div[2]/div/div[2]/div[2]/div[1]/div[1]/div/div/div[1]/div[3]/div[1]/span"))).text
#     placeholderInput = WebDriverWait(browser, 20).until(EC.visibility_of_element_located((By.XPATH, "/html/body/div[2]/div/div[4]/div[2]/div[2]/div/div[2]/div[2]/div[1]/div[1]/div/div/div[1]/div[3]/div[2]/div/div/span"))).text
#     try:
#         assert placeholderInput == placeholderLabel, "Label Legal name tidak sama dengan placeholder!"
#     finally:
#         print("Label Legal name sama dengan placeholder!")

# def pos001():
#     elementLoginPresent()
#     login()
#     headerSidebar()
#     findSystemProvider()
#     findSupervisionalBE()
#     findOperationalBE()

# def pos025():
#     elementLoginPresent()
#     login()
#     headerSidebar()
#     findSupervisionalBE()
#     clickSupervisionalBE()
#     tombolAddSupervisionalBE()
#     formParent()
#     formLegalName()
#     formBrandName()
#     formPhoneNumber()
#     formEmail()
#     formWebAddress()
#     formAddressHome()
#     formTaxCode()
#     usernameCredentials()
#     passwordCredentials()
#     confirmPasswordCredentials()
#     contractStart()
#     contractEnd()
#     SaveAddSupervisionalBE()

# def pos025a():
#     elementLoginPresent()
#     login()
#     headerSidebar()
#     findSupervisionalBE()
#     clickSupervisionalBE()
#     tombolAddSupervisionalBE()
#     formParent()
#     formLegalName()
#     formBrandName()
#     formPhoneNumber()
#     formEmail()
#     formWebAddress()
#     formAddressHome()
#     formTaxCode()
#     usernameCredentials()
#     passwordCredentials()
#     confirmPasswordCredentials()
#     contractStart()
#     SaveAddSupervisionalBE()

# def pos026():
#     elementLoginPresent()
#     login()
#     headerSidebar()
#     findSupervisionalBE()
#     clickSupervisionalBE()
#     tombolAddSupervisionalBE()
#     buttonLanguage()

# def pos027():
#     elementLoginPresent()
#     login()
#     headerSidebar()
#     findSupervisionalBE()
#     clickSupervisionalBE()
#     tombolAddSupervisionalBE()
#     formParent()

# def pos028():
#     elementLoginPresent()
#     login()
#     headerSidebar()
#     findSupervisionalBE()
#     clickSupervisionalBE()
#     tombolAddSupervisionalBE()
#     moveTab()

# def pos029():
#     elementLoginPresent()
#     login()
#     headerSidebar()
#     findSupervisionalBE()
#     clickSupervisionalBE()
#     tombolAddSupervisionalBE()
#     formPhoneNumber()

# def pos030():
#     elementLoginPresent()
#     login()
#     headerSidebar()
#     findSupervisionalBE()
#     clickSupervisionalBE()
#     tombolAddSupervisionalBE()
#     formEmail()

# def pos031():
#     elementLoginPresent()
#     login()
#     headerSidebar()
#     findSupervisionalBE()
#     clickSupervisionalBE()
#     tombolAddSupervisionalBE()
#     passwordCredentials()

# def pos032():
#     elementLoginPresent()
#     login()
#     headerSidebar()
#     findSupervisionalBE()
#     clickSupervisionalBE()
#     tombolAddSupervisionalBE()
#     passwordCredentials()
#     confirmPasswordCredentials()

# def pos033():
#     elementLoginPresent()
#     login()
#     headerSidebar()
#     findSupervisionalBE()
#     clickSupervisionalBE()
#     tombolAddSupervisionalBE()
#     passwordCredentials()
#     eyePasswordCredentials()

# def pos035():
#     elementLoginPresent()
#     login()
#     headerSidebar()
#     findSupervisionalBE()
#     clickSupervisionalBE()
#     tombolAddSupervisionalBE()
#     formParent()
#     formLegalName()
#     formBrandName()
#     formPhoneNumber()
#     formEmail()
#     formWebAddress()
#     formAddressHome()
#     formTaxCode()
#     usernameCredentials()
#     passwordCredentials()
#     confirmPasswordCredentials()
#     contractStart()
#     SaveAddSupervisionalBE()

# def pos036():
#     elementLoginPresent()
#     login()
#     headerSidebar()
#     findSupervisionalBE()
#     clickSupervisionalBE()
#     tombolAddSupervisionalBE()
#     formParent()
#     formLegalName()
#     formBrandName()
#     formPhoneNumber()
#     formEmail()
#     formWebAddress()
#     formAddressHome()
#     formTaxCode()
#     usernameCredentials()
#     passwordCredentials()
#     confirmPasswordCredentials()
#     contractEnd()
#     SaveAddSupervisionalBE()

# def pos037():
#     elementLoginPresent()
#     login()
#     headerSidebar()
#     findSupervisionalBE()
#     clickSupervisionalBE()
#     tombolAddSupervisionalBE()
#     formWebAddress()

# def pos038():
#     elementLoginPresent()
#     login()
#     headerSidebar()
#     findSupervisionalBE()
#     clickSupervisionalBE()
#     buttonEditSupervisionalBE()
#     formEditParent()
#     formEditLegalName()
#     formEditBrandName()
#     formEditPhoneNumber()
#     formEditEmailAddress()
#     formEditWebAddress()
#     formEditTaxCode()
#     formEditUsernameCredentials()
#     editContractStart()
#     editContractEnd()
#     SaveAddSupervisionalBE()

# def pos039():
#     elementLoginPresent()
#     login()
#     headerSidebar()
#     findSupervisionalBE()
#     clickSupervisionalBE()
#     buttonEditSupervisionalBE()
#     formParent()

# def pos040():
#     elementLoginPresent()
#     login()
#     headerSidebar()
#     findSupervisionalBE()
#     clickSupervisionalBE()
#     buttonEditSupervisionalBE()
#     buttonLanguage()

# def pos041():
#     elementLoginPresent()
#     login()
#     headerSidebar()
#     findSupervisionalBE()
#     clickSupervisionalBE()
#     buttonEditSupervisionalBE()
#     moveTab()

# def pos042():
#     elementLoginPresent()
#     login()
#     headerSidebar()
#     findSupervisionalBE()
#     clickSupervisionalBE()
#     buttonEditSupervisionalBE()
#     formEditPhoneNumber()

# def pos043():
#     elementLoginPresent()
#     login()
#     headerSidebar()
#     findSupervisionalBE()
#     clickSupervisionalBE()
#     buttonEditSupervisionalBE()
#     formEditEmailAddress()

# def pos044():
#     elementLoginPresent()
#     login()
#     headerSidebar()
#     findSupervisionalBE()
#     clickSupervisionalBE()
#     buttonEditSupervisionalBE()
#     formEditParent()
#     formEditLegalName()
#     formEditBrandName()
#     formEditPhoneNumber()
#     formEditEmailAddress()
#     formEditWebAddress()
#     formEditTaxCode()
#     formEditUsernameCredentials()
#     editContractStart()
#     SaveAddSupervisionalBE()

# def pos045():
#     elementLoginPresent()
#     login()
#     headerSidebar()
#     findSupervisionalBE()
#     clickSupervisionalBE()
#     buttonEditSupervisionalBE()
#     formEditParent()
#     formEditLegalName()
#     formEditBrandName()
#     formEditPhoneNumber()
#     formEditEmailAddress()
#     formEditWebAddress()
#     formEditTaxCode()
#     formEditUsernameCredentials()
#     editContractEnd()
#     SaveAddSupervisionalBE()

# def pos046():
#     elementLoginPresent()
#     login()
#     headerSidebar()
#     findSupervisionalBE()
#     clickSupervisionalBE()
#     buttonEditSupervisionalBE()
#     formEditWebAddress()

# def pos046a():
#     elementLoginPresent()
#     login()
#     headerSidebar()
#     findSupervisionalBE()
#     clickSupervisionalBE()
#     buttonEditSupervisionalBE()
#     formEditParent()
#     formEditLegalName()
#     formEditBrandName()
#     formEditPhoneNumber()
#     formEditEmailAddress()
#     formEditWebAddress()
#     formEditTaxCode()
#     formEditUsernameCredentials()
#     SaveAddSupervisionalBE()

# def pos047():
#     elementLoginPresent()
#     login()
#     headerSidebar()
#     findSupervisionalBE()
#     clickSupervisionalBE()
#     buttonViewSupervisionalBE()
#     buttonLanguage()

# def pos047A():
#     elementLoginPresent()
#     login()
#     headerSidebar()
#     findSupervisionalBE()
#     clickSupervisionalBE()
#     buttonViewSupervisionalBE()
#     buttonBackToListSupervisionalBE()

# def pos047B():
#     elementLoginPresent()
#     login()
#     headerSidebar()
#     findSupervisionalBE()
#     clickSupervisionalBE()
#     buttonViewSupervisionalBE()

# def pos048():
#     elementLoginPresent()
#     login()
#     headerSidebar()
#     findSupervisionalBE()
#     clickSupervisionalBE()
#     buttondeleteSupervisionalBE()
#     buttonAcceptDeleteSupervisionalBE()

# def pos049():
#     elementLoginPresent()
#     login()
#     headerSidebar()
#     findSupervisionalBE()
#     clickSupervisionalBE()
#     buttonBreadCrumbHome()

# def pos049a():
#     elementLoginPresent()
#     login()
#     headerSidebar()
#     findSupervisionalBE()
#     clickSupervisionalBE()
#     formParentfilterInputParent()
#     buttonFilter()


# #kosongkan semua isi field
# def neg013():
#     elementLoginPresent()
#     login()
#     headerSidebar()
#     findSupervisionalBE()
#     clickSupervisionalBE()
#     tombolAddSupervisionalBE()
#     formParent()
#     formLegalName()
#     formBrandName()
#     formPhoneNumber()
#     formEmail()
#     formWebAddress()
#     formAddressHome()
#     formTaxCode()
#     usernameCredentials()
#     passwordCredentials()
#     confirmPasswordCredentials()
#     contractStart()
#     contractEnd()
#     SaveAddSupervisionalBE()

# #fill the data with more than 300 character
# def neg014():
#     elementLoginPresent()
#     login()
#     headerSidebar()
#     findSupervisionalBE()
#     clickSupervisionalBE()
#     tombolAddSupervisionalBE()
#     formParent()
#     formLegalName()
#     formBrandName()
#     formPhoneNumber()
#     formEmail()
#     formWebAddress()
#     formAddressHome()
#     formTaxCode()
#     usernameCredentials()
#     passwordCredentials()
#     confirmPasswordCredentials()
#     contractStart()
#     contractEnd()
#     SaveAddSupervisionalBE()

# #fill the data with more than 300 character and then see the error validation
# def neg015():
#     elementLoginPresent()
#     login()
#     headerSidebar()
#     findSupervisionalBE()
#     clickSupervisionalBE()
#     tombolAddSupervisionalBE()
#     formParent()
#     formLegalName()
#     formBrandName()
#     formPhoneNumber()
#     formEmail()
#     formWebAddress()
#     formAddressHome()
#     formTaxCode()
#     usernameCredentials()
#     passwordCredentials()
#     confirmPasswordCredentials()
#     contractStart()
#     contractEnd()
#     SaveAddSupervisionalBE()

# #fill other field and fill the username field with space
# def neg016():
#     elementLoginPresent()
#     login()
#     headerSidebar()
#     findSupervisionalBE()
#     clickSupervisionalBE()
#     tombolAddSupervisionalBE()
#     formParent()
#     formLegalName()
#     formBrandName()
#     formPhoneNumber()
#     formEmail()
#     formWebAddress()
#     formAddressHome()
#     formTaxCode()
#     usernameCredentials()
#     passwordCredentials()
#     confirmPasswordCredentials()
#     contractStart()
#     contractEnd()
#     SaveAddSupervisionalBE()

# #fill the password with the confirmation password with other character
# def neg017():
#     elementLoginPresent()
#     login()
#     headerSidebar()
#     findSupervisionalBE()
#     clickSupervisionalBE()
#     tombolAddSupervisionalBE()
#     formParent()
#     formLegalName()
#     formBrandName()
#     formPhoneNumber()
#     formEmail()
#     formWebAddress()
#     formAddressHome()
#     formTaxCode()
#     usernameCredentials()
#     passwordCredentials()
#     confirmPasswordCredentials()
#     contractStart()
#     contractEnd()
#     SaveAddSupervisionalBE()

# #fill the start day with today and end day with yesterday
# def neg018():
#     elementLoginPresent()
#     login()
#     headerSidebar()
#     findSupervisionalBE()
#     clickSupervisionalBE()
#     tombolAddSupervisionalBE()
#     formParent()
#     formLegalName()
#     formBrandName()
#     formPhoneNumber()
#     formEmail()
#     formWebAddress()
#     formAddressHome()
#     formTaxCode()
#     usernameCredentials()
#     passwordCredentials()
#     confirmPasswordCredentials()
#     contractStart()
#     contractEnd()
#     SaveAddSupervisionalBE()

# #add new supervisional BE with data that was existing
# def neg019():
#     elementLoginPresent()
#     login()
#     headerSidebar()
#     findSupervisionalBE()
#     clickSupervisionalBE()
#     tombolAddSupervisionalBE()
#     formParent()
#     formLegalName()
#     formBrandName()
#     formPhoneNumber()
#     formEmail()
#     formWebAddress()
#     formAddressHome()
#     formTaxCode()
#     usernameCredentials()
#     passwordCredentials()
#     confirmPasswordCredentials()
#     contractStart()
#     contractEnd()
#     SaveAddSupervisionalBE()

# #empty required fields
# def neg020():
#     elementLoginPresent()
#     login()
#     headerSidebar()
#     findSupervisionalBE()
#     clickSupervisionalBE()
#     tombolAddSupervisionalBE()
#     formParent()
#     formLegalName()
#     formBrandName()
#     formPhoneNumber()
#     formEmail()
#     formWebAddress()
#     formAddressHome()
#     formTaxCode()
#     usernameCredentials()
#     passwordCredentials()
#     confirmPasswordCredentials()
#     contractStart()
#     contractEnd()
#     SaveAddSupervisionalBE()

# #edit user and save with existing user
# def neg021():
#     elementLoginPresent()
#     login()
#     headerSidebar()
#     findSupervisionalBE()
#     clickSupervisionalBE()
#     buttonEditSupervisionalBE()
#     formEditParent()
#     formEditLegalName()
#     formEditBrandName()
#     formEditPhoneNumber()
#     formEditEmailAddress()
#     formEditWebAddress()
#     formEditTaxCode()
#     formEditUsernameCredentials()
#     editContractStart()
#     SaveAddSupervisionalBE()

# #require fields to be empty
# def neg022():
#     elementLoginPresent()
#     login()
#     headerSidebar()
#     findSupervisionalBE()
#     clickSupervisionalBE()
#     buttonEditSupervisionalBE()
#     formEditParent()
#     formEditLegalName()
#     formEditBrandName()
#     formEditPhoneNumber()
#     formEditEmailAddress()
#     formEditWebAddress()
#     formEditTaxCode()
#     formEditUsernameCredentials()
#     editContractStart()
#     SaveAddSupervisionalBE()

# #verify cross mark change into warning after error validation page was showing
# def neg023():
#     elementLoginPresent()
#     login()
#     headerSidebar()
#     findSupervisionalBE()
#     clickSupervisionalBE()
#     buttonEditSupervisionalBE()
#     formEditParent()
#     formEditLegalName()
#     formEditBrandName()
#     formEditPhoneNumber()
#     formEditEmailAddress()
#     formEditWebAddress()
#     formEditTaxCode()
#     formEditUsernameCredentials()
#     editContractStart()
#     SaveAddSupervisionalBE()

# #error validation if username is filled with space in edit
# def pos024():
#     elementLoginPresent()
#     login()
#     headerSidebar()
#     findSupervisionalBE()
#     clickSupervisionalBE()
#     buttonEditSupervisionalBE()
#     formEditParent()
#     formEditLegalName()
#     formEditBrandName()
#     formEditPhoneNumber()
#     formEditEmailAddress()
#     formEditWebAddress()
#     formEditTaxCode()
#     formEditUsernameCredentials()
#     editContractStart()
#     SaveAddSupervisionalBE()

# #connection loss when add data
# def abn003():
#     elementLoginPresent()
#     login()
#     headerSidebar()
#     findSupervisionalBE()
#     clickSupervisionalBE()
#     tombolAddSupervisionalBE()
#     formParent()
#     formLegalName()
#     formBrandName()
#     formPhoneNumber()
#     formEmail()
#     formWebAddress()
#     formAddressHome()
#     formTaxCode()
#     usernameCredentials()
#     passwordCredentials()
#     confirmPasswordCredentials()
#     contractStart()
#     contractEnd()
#     SaveAddSupervisionalBE()

# #connection loss when update data in edit
# def abn004():
#     elementLoginPresent()
#     login()
#     headerSidebar()
#     findSupervisionalBE()
#     clickSupervisionalBE()
#     buttonEditSupervisionalBE()
#     formEditParent()
#     formEditLegalName()
#     formEditBrandName()
#     formEditPhoneNumber()
#     formEditEmailAddress()
#     formEditWebAddress()
#     formEditTaxCode()
#     formEditUsernameCredentials()
#     editContractStart()
#     SaveAddSupervisionalBE()

# def pos034():
#     elementLoginPresent()
#     login()
#     headerSidebar()
#     findSupervisionalBE()
#     clickSupervisionalBE()
#     tombolAddSupervisionalBE()
#     assertField()

# pos001()
# pos025()
# pos025a()
# pos026()
# pos027()
# pos028()
# pos029()
# pos030()
# pos031()
# pos032()
# pos033()
# pos034()
# pos035()
# pos036()
# pos037()
# pos038()
# pos039()
# pos040()
# pos041()
# pos042()
# pos043()
# pos044()
# pos045()
# pos046()
# pos046a()
# pos047()
# pos047A()
# pos047B()
# pos048()
# pos049()
# pos049a()
